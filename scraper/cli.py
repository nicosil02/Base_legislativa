"""CLI del scraper.

Uso:
  python -m scraper.cli init                       # crea DB y baja comisiones
  python -m scraper.cli update [--full] [--limit N]
  python -m scraper.cli export [--out proyectos.json]
  python -m scraper.cli query --comision 4
  python -m scraper.cli query --estado "EN COMISIÓN"
  python -m scraper.cli show 14515
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from scraper.api import ApiClient
from scraper.db import Database
from scraper.export import export_json
from scraper.sync import PER_PAR_ID_ACTUAL, env_max_proyectos, run_sync

DEFAULT_DB = "proyectos.db"
DEFAULT_JSON = "proyectos.json"


def cmd_init(args) -> int:
    with Database(args.db) as db:
        db.init_schema()
        if db.count_comisiones() == 0:
            client = ApiClient()
            comis = client.list_comisiones()
            db.upsert_comisiones(comis)
            print(f"DB inicializada en {args.db} — {len(comis)} comisiones cargadas.")
        else:
            print(f"DB ya tenía {db.count_comisiones()} comisiones; esquema asegurado.")
    return 0


def cmd_update(args) -> int:
    with Database(args.db) as db:
        db.init_schema()
        max_p = args.limit if args.limit is not None else env_max_proyectos()
        stats = run_sync(db, full=args.full, max_proyectos=max_p)
        print(
            f"Sync terminado: vistos={stats.vistos} nuevos={stats.nuevos} "
            f"actualizados={stats.actualizados} detail_fetches={stats.detail_fetches} "
            f"errores={stats.errores}"
        )
    return 0


def cmd_export(args) -> int:
    with Database(args.db) as db:
        n = export_json(db, args.out)
        print(f"Exportados {n} proyectos a {args.out}")
    return 0


def cmd_query(args) -> int:
    with Database(args.db) as db:
        sql = (
            "SELECT p.per_par_id, p.pley_num, p.proyecto_ley, p.estado, p.fec_presentacion, "
            "       p.proponente, p.grupo_parlamentario, p.tema, p.tema_manual, "
            "       (SELECT GROUP_CONCAT(pc.nombre, ' | ') FROM proyecto_comision pc "
            "        WHERE pc.per_par_id=p.per_par_id AND pc.pley_num=p.pley_num) AS comisiones, "
            "       p.titulo "
            "FROM proyectos p"
        )
        clauses: list[str] = []
        params: list = []
        if args.comision is not None:
            clauses.append(
                "EXISTS (SELECT 1 FROM proyecto_comision pc "
                "WHERE pc.per_par_id=p.per_par_id AND pc.pley_num=p.pley_num AND pc.comision_id=?)"
            )
            params.append(args.comision)
        if args.tema:
            clauses.append("p.tema = ?")
            params.append(args.tema)
        if args.estado:
            clauses.append("p.estado = ?")
            params.append(args.estado)
        if args.proponente:
            clauses.append("p.proponente = ?")
            params.append(args.proponente)
        if args.partido:
            clauses.append("p.grupo_parlamentario = ?")
            params.append(args.partido)
        if clauses:
            sql += " WHERE " + " AND ".join(clauses)
        sql += " ORDER BY p.fec_presentacion DESC LIMIT ?"
        params.append(args.limit)
        rows = db.conn.execute(sql, params).fetchall()
        for r in rows:
            tema_marker = "*" if r["tema_manual"] else " "
            print(
                f"{r['proyecto_ley']:18s} {r['estado']:22s} {r['fec_presentacion'][:10]} "
                f"{(r['proponente'] or '-'):12s} {(r['grupo_parlamentario'] or '-'):20s} "
                f"{tema_marker}<{r['tema'] or '-'}> "
                f"[{r['comisiones'] or '-'}] {r['titulo'][:60]}"
            )
        print(f"\n{len(rows)} resultado(s)")
    return 0


def cmd_recategorizar(args) -> int:
    """Re-clasifica con el clasificador automático SOLO los proyectos sin
    etiqueta manual. Los marcados como tema_manual=1 (importados del Excel)
    no se tocan."""
    with Database(args.db) as db:
        db.init_schema()
        if args.force:
            with db.tx() as c:
                c.execute("UPDATE proyectos SET tema=NULL, tema_manual=0")
            print("--force: limpié todas las etiquetas, incluyendo manuales.")
        rows = db.conn.execute(
            "SELECT per_par_id, pley_num, titulo, sumilla FROM proyectos "
            "WHERE tema_manual = 0"
        ).fetchall()
        total = len(rows)
        print(f"Re-clasificando {total} proyectos (los manuales se respetan)...")
        for i, r in enumerate(rows, 1):
            db.classify_and_save(r["per_par_id"], r["pley_num"], r["titulo"], r["sumilla"])
            if i % 1000 == 0:
                print(f"  {i}/{total}")
        print(f"Listo: {total} proyectos re-clasificados.")
        print("\nDistribución por tema:")
        for r in db.conn.execute(
            "SELECT tema, SUM(tema_manual) AS manuales, COUNT(*) AS total "
            "FROM proyectos GROUP BY tema ORDER BY total DESC"
        ):
            print(f"  {(r['tema'] or '(sin tema)'):36s} total={r['total']:5d}  manuales={r['manuales']}")
    return 0


def cmd_importar_temas(args) -> int:
    """Importa los temas etiquetados a mano desde un Excel.

    Se esperan columnas 'PL' (número entero) y 'Tema'. Cada match se guarda
    con tema_manual=1 para que el clasificador no lo sobrescriba.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl no instalado. Corre: pip install openpyxl", file=sys.stderr)
        return 2
    print(f"Leyendo {args.excel}...")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.rows)]
    if "PL" not in headers or "Tema" not in headers:
        print(f"ERROR: el Excel necesita columnas 'PL' y 'Tema'. Encontré: {headers}", file=sys.stderr)
        return 2
    i_pl = headers.index("PL")
    i_tema = headers.index("Tema")

    with Database(args.db) as db:
        db.init_schema()
        n_match, n_unknown, n_skip = 0, 0, 0
        skipped_temas: set[str] = set()
        valid = set(db.conn.execute("SELECT pley_num FROM proyectos").fetchall())
        valid = {r[0] for r in valid}
        for row in ws.iter_rows(min_row=2, values_only=True):
            pl, tema = row[i_pl], row[i_tema]
            if pl is None or tema is None:
                n_skip += 1
                continue
            try:
                pl = int(pl)
            except (ValueError, TypeError):
                n_skip += 1
                continue
            if pl not in valid:
                n_unknown += 1
                continue
            db.set_tema(args.per_par_id, pl, str(tema).strip(), manual=True)
            n_match += 1
            if tema not in (
                "Educación","Trabajo","Salud","Tributos","Banca","Pensiones",
                "Ambiente","Agricultura","Horeca","Transporte","Construcción",
                "Transporte y telecomunicaciones","Energía y minas","Pesca","Energía",
                "Minería","Telecomunicaciones","Comercio","Saneamiento",
                "Control de la actividad privada","Infraestructura","Mype",
                "Inmobiliario","Informalidad","Consumo masivo","Seguros","Deporte","Otros",
            ):
                skipped_temas.add(str(tema))
        print(f"Importados: {n_match}  |  PLs no encontrados en DB: {n_unknown}  |  Filas vacías: {n_skip}")
        if skipped_temas:
            print(f"  Temas no estándar encontrados: {sorted(skipped_temas)}")
        print("\nDistribución por tema (post-import):")
        for r in db.conn.execute(
            "SELECT tema, SUM(tema_manual) AS manuales, COUNT(*) AS total "
            "FROM proyectos GROUP BY tema ORDER BY total DESC"
        ):
            print(f"  {(r['tema'] or '(sin tema)'):36s} total={r['total']:5d}  manuales={r['manuales']}")
    return 0




def cmd_show(args) -> int:
    with Database(args.db) as db:
        row = db.conn.execute(
            "SELECT * FROM proyectos WHERE per_par_id=? AND pley_num=?",
            (args.per_par_id, args.pley_num),
        ).fetchone()
        if not row:
            print(f"Proyecto {args.per_par_id}/{args.pley_num} no encontrado.", file=sys.stderr)
            return 1
        print(f"Proyecto de Ley: {row['proyecto_ley']}")
        print(f"Título: {row['titulo']}")
        print(f"Estado: {row['estado']}  (id={row['estado_id']})")
        print(f"Presentado: {row['fec_presentacion']}")
        print(f"Proponente: {row['proponente']}  Grupo: {row['grupo_parlamentario']}")
        print(f"Portal: {row['url_portal']}")
        if row["url_pdf"]:
            print(f"PDF: {row['url_pdf']}")
        if row["sumilla"]:
            print(f"\nSumilla:\n{row['sumilla']}")
        coms = db.conn.execute(
            "SELECT nombre FROM proyecto_comision WHERE per_par_id=? AND pley_num=?",
            (args.per_par_id, args.pley_num),
        ).fetchall()
        if coms:
            print("\nComisiones: " + ", ".join(c["nombre"] for c in coms))
        if row["tema"]:
            origen = "manual (Excel)" if row["tema_manual"] else "auto"
            print(f"Tema: {row['tema']}  [{origen}]")
        segs = db.conn.execute(
            "SELECT fecha, estado, comisiones, observacion FROM seguimientos "
            "WHERE per_par_id=? AND pley_num=? ORDER BY fecha DESC",
            (args.per_par_id, args.pley_num),
        ).fetchall()
        if segs:
            print("\nHistorial:")
            for s in segs:
                fecha = (s["fecha"] or "")[:10]
                line = f"  {fecha}  {s['estado']}"
                if s["comisiones"]:
                    line += f"  ({s['comisiones']})"
                print(line)
    return 0


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="scraper", description="Scraper de proyectos de ley del Congreso del Perú")
    p.add_argument("--db", default=DEFAULT_DB, help=f"Ruta del SQLite (default: {DEFAULT_DB})")
    p.add_argument("-v", "--verbose", action="store_true")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("init", help="crea DB y carga comisiones").set_defaults(func=cmd_init)

    up = sub.add_parser("update", help="corre sync incremental")
    up.add_argument("--full", action="store_true", help="re-enriquece todos los proyectos (no solo los cambiados)")
    up.add_argument("--limit", type=int, default=None, help="máximo de proyectos a procesar")
    up.set_defaults(func=cmd_update)

    ex = sub.add_parser("export", help="exporta a JSON")
    ex.add_argument("--out", default=DEFAULT_JSON)
    ex.set_defaults(func=cmd_export)

    q = sub.add_parser("query", help="lista proyectos con filtros")
    q.add_argument("--comision", type=int)
    q.add_argument("--estado")
    q.add_argument("--proponente", help="ej. Congreso, Ejecutivo, Ciudadanos, Regional")
    q.add_argument("--partido", help="grupo parlamentario exacto, ej. 'Perú Libre'")
    q.add_argument("--tema", help="categoría temática, ej. 'Tecnología', 'Agricultura', 'Farma'")
    q.add_argument("--limit", type=int, default=50)
    q.set_defaults(func=cmd_query)

    rc = sub.add_parser("recategorizar", help="re-aplica el clasificador automático a los PLs sin etiqueta manual")
    rc.add_argument("--force", action="store_true",
                    help="limpia TODAS las etiquetas (incluyendo manuales) y vuelve a correr el clasificador")
    rc.set_defaults(func=cmd_recategorizar)

    it = sub.add_parser("importar-temas", help="carga temas etiquetados a mano desde un Excel (tema_manual=1)")
    it.add_argument("excel", help="ruta al archivo .xlsx con columnas 'PL' y 'Tema'")
    it.add_argument("--sheet", help="nombre de la hoja (por defecto la primera)")
    it.add_argument("--per-par-id", dest="per_par_id", type=int, default=PER_PAR_ID_ACTUAL)
    it.set_defaults(func=cmd_importar_temas)

    sh = sub.add_parser("show", help="muestra un proyecto e historial")
    sh.add_argument("pley_num", type=int)
    sh.add_argument("--per-par-id", dest="per_par_id", type=int, default=PER_PAR_ID_ACTUAL)
    sh.set_defaults(func=cmd_show)

    return p


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s — %(message)s",
    )
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
